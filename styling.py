from openpyxl.styles import Font, Alignment
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string


class Styling:
    """
    A class representing an Excel styling object.

    """

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.wb = load_workbook(self.xlsx_path)

    def bold_headers(self, sheet_name: str):
        """
        Make headers bold in the specified Excel sheet.

        :param sheet_name: The name of the sheet in the Excel workbook.
        """
        for cell in self.wb[sheet_name][1]:
            cell.font = Font(bold=True)

        self.wb.save(self.xlsx_path)

    def apply_autofit_columns(self, sheet_name: str):
        """
        Autofits sheet columns to fit the longest text in the column.

        :param sheet_name: The name of the sheet in the Excel workbook.
        """
        for column_cells in self.wb[sheet_name].columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                self.wb[sheet_name].column_dimensions[new_column_letter].width = new_column_length * 1.25

        self.wb.save(self.xlsx_path)

    def apply_auto_filter_headers(self, sheet_name: str):
        # Add filter to the top row (headers)
        self.wb[sheet_name].auto_filter.ref = self.wb[sheet_name].dimensions
        self.wb.save(self.xlsx_path)

    def apply_basic_styling(self, sheet_name: str):
        """
        Apply basic styling to the specified sheet.

        :param sheet_name: The name of the sheet to which styling will be applied.
        """
        self.bold_headers(sheet_name)
        self.apply_auto_filter_headers(sheet_name)
        self.apply_autofit_columns(sheet_name)

    def color_row(self, sheet_name: str, color: str, start_row: int, end_row: int):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in self.wb[sheet_name].iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                cell.fill = fill
        self.wb.save(self.xlsx_path)

    def color_column(self, sheet_name: str, color: str, column: int):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in self.wb[sheet_name].iter_rows(min_row=2, min_col=column, max_col=column):
            row[0].fill = fill
        self.wb.save(self.xlsx_path)
