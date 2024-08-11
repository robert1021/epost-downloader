from openpyxl import Workbook


class XlsxBuilder:

    def __init__(self):
        self.workbook = Workbook()

    def add_sheet(self, sheet_name: str, row_data: list[list]):
        sheet = self.workbook.create_sheet(sheet_name)
        for row in row_data:
            sheet.append(row)

    def save(self, filename):
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        self.workbook.save(filename)
